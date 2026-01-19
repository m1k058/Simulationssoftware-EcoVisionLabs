import openpyxl
import pandas as pd

# Dateiname anpassen
input_file = 'Erweiterte Speicherlogik.xlsx' 
output_file = 'Excel_Logik_Extrakt.txt'

# Excel laden - WICHTIG: data_only=False lädt die Formel, nicht den Wert
wb = openpyxl.load_workbook(input_file, data_only=False)

with open(output_file, 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        f.write(f"\n--- SHEET: {sheet_name} ---\n")
        
        # Iteriere durch die genutzten Zeilen (z.B. erste 5 Zeilen reichen oft für die Logik)
        # Wir nehmen hier Zeile 1 (Header) und Zeile 2 (erste Formel-Zeile)
        rows_to_scan = [1, 2] 
        
        for r_idx in rows_to_scan:
            row_data = []
            for c_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                value = cell.value
                
                # Wenn es eine Formel ist (beginnt mit =), speichern wir sie als String
                if isinstance(value, str) and value.startswith('='):
                    val_str = f"FORMEL: {value}"
                else:
                    val_str = str(value)
                
                row_data.append(val_str)
            
            # Schreiben als Pipe-getrennte Linie für gute Lesbarkeit
            f.write(f"Zeile {r_idx}: " + " | ".join(row_data) + "\n")

print(f"Logik erfolgreich in {output_file} extrahiert.")